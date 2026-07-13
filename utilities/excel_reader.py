from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def get_login_data():
        workbook = load_workbook("testdata/login_data.xlsx")
        sheet = workbook.active

        data = []

        # Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            username, password = row
            data.append((username, password))

        workbook.close()

        return data